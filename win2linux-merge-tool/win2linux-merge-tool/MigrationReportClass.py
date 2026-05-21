import csv
from datetime import datetime

import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import os

from utils.utils import get_reports_path

try:
    import config_loader
except ImportError:
    config_loader = None


def _reports_enabled():
    """Return True if migration reports should be written (setting is enabled)."""
    if config_loader is None or not getattr(config_loader, 'settings', None):
        return True  # Default to enabled when config not loaded
    return config_loader.settings.get('write_migration_reports', True)


class MigrationReport:

    def __init__(self, title, description=""):
        self.title = title
        self.content = []
        self.description = ""

    def add_report_entry(self, report_entry):

        self.content.append(report_entry)

    def get_errors_count(self):
        """
        Returns the number of entries in the report.
        """
        return len(self.content)

    def save_report_to_csv(self, path_to_csv_file="reports/migration_report.csv"):
        """
        Saves the migration report to a CSV file.

        Args:
            path_to_csv_file (str): The path to the CSV file where the report should be saved.
                Defaults to "reports/migration_report.csv". Relative paths are resolved to a writable reports directory.
        """
        if not _reports_enabled():
            return
        path_to_csv_file = get_reports_path(path_to_csv_file)
        # Create the 'reports' directory if it doesn't exist
        os.makedirs(os.path.dirname(path_to_csv_file), exist_ok=True)
        # Get the current date and time in a suitable format
        now = datetime.now()
        timestamp = now.strftime("%Y%m%d_%H%M%S")

        # Construct the new filename with the timestamp
        base, ext = os.path.splitext(path_to_csv_file)
        new_filename = f"{base}_{timestamp}{ext}"
        try:
            with open(new_filename, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)

                # Write the header row
                if self.content:  # Check if there's any content to get headers from
                    header = self.content[0].keys()
                    writer.writerow(header)

                # Write the data rows
                for row in self.content:
                    writer.writerow(row.values())
        except Exception as e:
            print(f"An error occurred while saving the report to CSV: {e}")

    def save_report_to_txt(self, path_to_txt_file="reports/migration_report.txt"):
        """
        Appends the migration report to a TXT file.  Each entry is written on a new line.

        Args:
            path_to_txt_file (str): The path to the TXT file.
                Defaults to "reports/migration_report.txt". Relative paths are resolved to a writable reports directory.
        """
        if not _reports_enabled():
            return
        path_to_txt_file = get_reports_path(path_to_txt_file)
        # Create the 'reports' directory if it doesn't exist
        os.makedirs(os.path.dirname(path_to_txt_file), exist_ok=True)

        # Get the current date and time in a suitable format
        now = datetime.now()
        timestamp = now.strftime("%Y%m%d_%H%M%S")

        # Construct the new filename with the timestamp
        base, ext = os.path.splitext(path_to_txt_file)
        new_filename = f"{base}_{timestamp}{ext}"

        try:
            with open(new_filename, 'a', encoding='utf-8') as txtfile:
                txtfile.write(f"{self.title}\n")  # Add the title of the report
                txtfile.write("\n")
                txtfile.write(f"{self.description}\n") # Add description to the report
                if self.content:
                    header = ", ".join(self.content[0].keys()) + "\n"
                    txtfile.write(header)
                for row in self.content:
                    row_str = ", ".join(f"{key}: {value}" for key, value in row.items()) + "\n"
                    txtfile.write(row_str)
                txtfile.write("\n")  # Add a newline after all entries are written

        except Exception as e:
            print(f"An error occurred while saving the report to TXT: {e}")

    def save_report_to_excel(self, path_to_excel_file="reports/migration_report.xlsx"):
        """
        Appends the migration report to an Excel file as a new sheet.

        Args:
            path_to_excel_file (str): The path to the Excel file.
                Defaults to "reports/migration_report.xlsx". Relative paths are resolved to a writable reports directory.
        """
        if not _reports_enabled():
            return
        path_to_excel_file = get_reports_path(path_to_excel_file)
        # Create the 'reports' directory if it doesn't exist
        os.makedirs(os.path.dirname(path_to_excel_file), exist_ok=True)
        # Get the current date and time in a suitable format
        now = datetime.now()
        timestamp = now.strftime("%Y%m%d_%H%M%S")

        # Construct the new filename with the timestamp
        base, ext = os.path.splitext(path_to_excel_file)
        new_filename = f"{base}_{timestamp}{ext}"
        try:
            # Create a Pandas DataFrame from the report content
            df = pd.DataFrame(self.content)

            # Load the workbook or create a new one if it doesn't exist
            try:
                workbook = openpyxl.load_workbook(new_filename)
            except FileNotFoundError:
                workbook = openpyxl.Workbook()

            # Sanitize the sheet name
            sheet_name = self.title[:31]  # Excel sheet names cannot exceed 31 characters
            if sheet_name in workbook.sheetnames:
                sheet_name = f"{sheet_name[:25]}_report"  # Truncate if needed and add a suffix

            # Create a new worksheet
            worksheet = workbook.create_sheet(title=sheet_name)

            # Append description to the worksheet
            worksheet['A1'] = self.description

            # Append the dataframe to the worksheet
            for row in dataframe_to_rows(df, header=True, index=False):
                worksheet.append(row)

            # Save the workbook
            workbook.save(new_filename)

        except Exception as e:
            print(f"An error occurred while saving the report to Excel: {e}")


# Example Usage
if __name__ == "__main__":
    report = MigrationReport("Data Migration Report")
    report.add_report_entry(
        {"source_table": "users", "target_table": "users_new", "status": "success", "records_migrated": 1000})
    report.add_report_entry({"source_table": "products", "target_table": "products_new", "status": "partial failure",
                             "records_migrated": 500, "errors": "Duplicate key error"})
    report.add_report_entry(
        {"source_table": "orders", "target_table": "orders_new", "status": "success", "records_migrated": 2000})

    print(f"Number of report entries: {report.get_errors_count()}") # Example usage of the new method

    report.save_report_to_csv()
    print("Report saved to reports/migration_report.csv")

    report.save_report_to_txt()
    print("Report appended to reports/migration_report.txt")

    report.save_report_to_excel()
    print("Report appended to reports/migration_report.xlsx")